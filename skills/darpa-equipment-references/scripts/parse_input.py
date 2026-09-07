#!/usr/bin/env python3
"""Parse DARPA equipment input from an xlsx Base (Year 1) tab OR a pasted-text file.

Usage:
    parse_input.py <input-path> [--out <output.json>]

Input can be:
  * .xlsx — reads "Base (Year 1)" sheet, equipment block starting at row 97
    (header) with rows 98..N until first blank row in column A.
  * any other file (.txt/.tsv/.md) — reads whitespace/tab-separated rows.

Output JSON shape:
    [
      {"no": 1, "description": "...", "vendor": "...", "qty": 1,
       "cost_per_item": 1114.0, "total_cost": 1114.0,
       "exclusive": "Yes", "competitive": "Yes", "supporting": "Yes",
       "link": "https://..."},
      ...
    ]
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path
from urllib.parse import urlparse, urlunparse

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)


def _http_ok(url: str, timeout: float = 10.0) -> bool:
    """Best-effort liveness check. Treat any 2xx/3xx response as OK."""
    if not url:
        return False
    try:
        req = urllib.request.Request(url, headers={"User-Agent": UA}, method="GET")
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return 200 <= r.status < 400
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, OSError):
        return False


def canonicalize_url(url: str, check: bool = True) -> tuple[str, str]:
    """Return (canonical_url, original_url).

    Canonical = same URL with query string and fragment stripped, BUT only if
    a live GET to that canonical URL succeeds (and `check` is True). Falls
    back to the original URL when the short form is unreachable.
    """
    if not url:
        return "", ""
    original = url.strip()
    parsed = urlparse(original)
    short = urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", "", ""))
    if short == original:
        return original, original
    if not check:
        return short, original
    if _http_ok(short):
        return short, original
    # Canonical doesn't resolve — keep original
    return original, original


SHEET_NAME = "Base (Year 1)"
EQUIPMENT_HEADER_ROW = 97  # 1-indexed; row immediately above first data row
COLUMNS = {
    "no": "A",
    "description": "B",
    "vendor": "C",
    "qty": "D",
    "cost_per_item": "E",
    "total_cost": "F",
    "exclusive": "G",
    "competitive": "I",
    "supporting": "L",
    "link": "N",
}


def _apply_url_canon(items: list[dict], check: bool) -> None:
    for it in items:
        canon, original = canonicalize_url(it.get("link", ""), check=check)
        it["link"] = canon
        it["link_original"] = original


def parse_xlsx(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("ERROR: openpyxl not installed. pip install openpyxl")

    wb = load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"ERROR: sheet {SHEET_NAME!r} not found. Available: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    items: list[dict] = []
    row = EQUIPMENT_HEADER_ROW + 1
    while True:
        no_cell = ws[f"{COLUMNS['no']}{row}"].value
        if no_cell is None or str(no_cell).strip() == "":
            break
        desc = ws[f"{COLUMNS['description']}{row}"].value
        if desc is None or str(desc).strip() == "":
            break

        qty_raw = ws[f"{COLUMNS['qty']}{row}"].value
        cost_raw = ws[f"{COLUMNS['cost_per_item']}{row}"].value
        total_raw = ws[f"{COLUMNS['total_cost']}{row}"].value

        try:
            qty = int(qty_raw) if qty_raw is not None else 1
        except (TypeError, ValueError):
            qty = 1
        try:
            cost = float(cost_raw) if cost_raw is not None else 0.0
        except (TypeError, ValueError):
            cost = 0.0
        try:
            total = float(total_raw) if total_raw is not None else cost * qty
        except (TypeError, ValueError):
            total = cost * qty
        if not total:
            total = cost * qty

        items.append({
            "no": int(no_cell) if isinstance(no_cell, (int, float)) else len(items) + 1,
            "description": str(desc).strip(),
            "vendor": str(ws[f"{COLUMNS['vendor']}{row}"].value or "").strip(),
            "qty": qty,
            "cost_per_item": round(cost, 2),
            "total_cost": round(total, 2),
            "exclusive": str(ws[f"{COLUMNS['exclusive']}{row}"].value or "").strip(),
            "competitive": str(ws[f"{COLUMNS['competitive']}{row}"].value or "").strip(),
            "supporting": str(ws[f"{COLUMNS['supporting']}{row}"].value or "").strip(),
            "link": str(ws[f"{COLUMNS['link']}{row}"].value or "").strip(),
        })
        row += 1
    return items


_money_re = re.compile(r"\$?\s*([\d,]+(?:\.\d+)?)")


def _money(s: str) -> float:
    if not s:
        return 0.0
    m = _money_re.search(s.replace(",", ""))
    if not m:
        try:
            return float(s)
        except ValueError:
            return 0.0
    return float(m.group(1))


def parse_paste(path: Path) -> list[dict]:
    """Parse a free-form pasted block. Each non-blank line is one item.

    Tries TAB split first, then runs-of-2+-spaces. Columns expected (in order):
        No  Description  Vendor  Qty  Cost  Total  Exclusive  Competitive  Supporting  Link

    Tolerant of missing trailing fields and re-numbers `no` to 1..N regardless
    of what the user pasted (so the row order in the file IS the EQ order).
    """
    text = path.read_text(encoding="utf-8", errors="replace")
    items: list[dict] = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        # Skip obvious header lines
        if re.match(r"^(no\.?|#|item)\b", line, re.I) and "description" in line.lower():
            continue

        if "\t" in line:
            cols = [c.strip() for c in line.split("\t")]
        else:
            cols = [c.strip() for c in re.split(r"\s{2,}", line)]
        cols = [c for c in cols if c]

        # Find URL — usually the last column
        link = ""
        for i in range(len(cols) - 1, -1, -1):
            if re.match(r"https?://", cols[i]):
                link = cols[i]
                cols = cols[:i] + cols[i + 1:]
                break

        if len(cols) < 2:
            continue

        # Pop the leading number if present (we re-number anyway)
        if cols and re.fullmatch(r"\d+", cols[0]):
            cols = cols[1:]

        description = cols[0] if len(cols) >= 1 else ""
        vendor = cols[1] if len(cols) >= 2 else ""

        # Try to find qty/cost/total in the remaining columns. Qty is the first
        # short pure-int after vendor; cost/total are the next $-money tokens.
        remaining = cols[2:]
        qty = 1
        cost = 0.0
        total = 0.0
        money_seen: list[float] = []
        for tok in remaining:
            if re.fullmatch(r"\d{1,3}", tok) and not money_seen:
                qty = int(tok)
                continue
            if "$" in tok or re.fullmatch(r"[\d,]+(\.\d+)?", tok):
                money_seen.append(_money(tok))
        if money_seen:
            cost = money_seen[0]
            total = money_seen[1] if len(money_seen) > 1 else cost * qty

        # Yes/Yes/Yes columns — pick out three trailing yes-ish tokens
        yes_tokens = [t for t in remaining if t.lower() in {"yes", "no", "y", "n"}]
        excl = yes_tokens[0].capitalize() if len(yes_tokens) >= 1 else "Yes"
        comp = yes_tokens[1].capitalize() if len(yes_tokens) >= 2 else "Yes"
        supp = yes_tokens[2].capitalize() if len(yes_tokens) >= 3 else "Yes"

        items.append({
            "no": len(items) + 1,
            "description": description,
            "vendor": vendor,
            "qty": qty,
            "cost_per_item": round(cost, 2),
            "total_cost": round(total or cost * qty, 2),
            "exclusive": excl,
            "competitive": comp,
            "supporting": supp,
            "link": link,
        })
    return items


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("input", type=Path)
    ap.add_argument("--out", type=Path, default=None)
    ap.add_argument("--no-check-urls", action="store_true",
                    help="Skip the canonical-URL liveness check (offline mode). "
                         "Still strips query/fragment from the link.")
    args = ap.parse_args()

    if not args.input.exists():
        sys.exit(f"ERROR: input not found: {args.input}")

    if args.input.suffix.lower() == ".xlsx":
        items = parse_xlsx(args.input)
    else:
        items = parse_paste(args.input)

    if not items:
        sys.exit("ERROR: no equipment items parsed from input")

    # Re-number 1..N to enforce that input order == EQ order
    for i, it in enumerate(items, 1):
        it["no"] = i

    _apply_url_canon(items, check=not args.no_check_urls)

    payload = json.dumps(items, indent=2, ensure_ascii=False)
    if args.out:
        args.out.write_text(payload, encoding="utf-8")
        print(f"Wrote {len(items)} items → {args.out}")
    else:
        print(payload)


if __name__ == "__main__":
    main()
