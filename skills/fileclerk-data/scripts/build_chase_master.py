#!/usr/bin/env python3
"""
build_chase_master.py — (re)build spreadsheet/chase-master.xlsx

One tab per Chase bank account. Chase activity exports are date-RANGE downloads
that overlap at their boundaries, so per account we STITCH the downloads on date
boundaries (newest-coverage download first, then older downloads fill only the
dates not already covered). This preserves each file's native newest-first order
and avoids double-counting — including genuine same-day duplicate transactions,
which a signature dedupe would wrongly merge.

Source files: every "<YYMMDD> Chase <last4> ALLREC.csv" in processed/.
Before running, file any raw "Chase<last4>_Activity_<YYYYMMDD>.csv" download into
processed/ under the ALLREC name (the renaming-finance-files Chase rule).

Integrity check: after stitching, the running Balance column must chain
(balance[newer] - amount[newer] == balance[older]) with zero breaks. A break
means a missing/duplicated transaction or a coverage gap.

Usage:  python build_chase_master.py
"""
import csv, glob, os, re
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill

BASE = "/Users/travis/Library/CloudStorage/OneDrive-izumanet/Finance/Invoice_processing"
OUT = f"{BASE}/spreadsheet/chase-master.xlsx"

# last4 -> (code, name). Extend as new accounts appear.
ACCT_MAP = {
    "7505": ("IN1", "Main Checking"),
    "3839": ("IN2", "Secondary Checking"),
    "3211": ("IS1", "Savings"),
    "6557": ("IT1", "Tech Checking"),
}
ORDER = ["7505", "3839", "3211", "6557"]

HEAD = ["Details", "Posting Date", "Description", "Amount", "Type",
        "Balance", "Check/Slip #", "source download"]
MONEY, DATEF = '#,##0.00', 'yyyy-mm-dd'


def acct(path):
    m = re.search(r'Chase[ _]?(\d{4})', os.path.basename(path), re.I)
    return m.group(1) if m else None


def load(path):
    out = []
    with open(path, newline='', encoding='utf-8-sig') as f:
        rd = csv.reader(f)
        next(rd)
        for row in rd:
            if not row or not row[0].strip():
                continue
            row = (row + [''] * 7)[:7]
            out.append(dict(
                details=row[0], date=datetime.strptime(row[1], "%m/%d/%Y"),
                desc=row[2],
                amt=float(row[3].replace(',', '')) if row[3].strip() else None,
                typ=row[4],
                bal=float(row[5].replace(',', '')) if row[5].strip() else None,
                chk=row[6], src=os.path.basename(path)))
    return out


def stitch(files):
    """Return (rows_newest_first, coverage_windows)."""
    loaded = [(f, load(f)) for f in files]
    # newest coverage first; on equal max-date, deepest history (earliest min-date) first
    loaded.sort(key=lambda t: (max(r['date'] for r in t[1]),
                               -min(r['date'] for r in t[1]).toordinal()), reverse=True)
    kept, min_incl, windows = [], None, []
    for f, rows in loaded:                       # rows already newest-first (Chase order)
        sel = rows if min_incl is None else [r for r in rows if r['date'] < min_incl]
        if sel:
            windows.append((os.path.basename(f),
                            min(r['date'] for r in sel).date(),
                            max(r['date'] for r in sel).date(), len(sel)))
            kept.extend(sel)                     # preserve native order; no re-sort
        # boundary = earliest date covered so far; may only move earlier
        file_min = min(r['date'] for r in rows)
        min_incl = file_min if min_incl is None else min(min_incl, file_min)
    return kept, windows


def qa_breaks(rows):
    breaks = []
    for j in range(1, len(rows)):
        nb, na, ob = rows[j - 1]['bal'], rows[j - 1]['amt'], rows[j]['bal']
        if None in (nb, na, ob):
            continue
        if abs((nb - na) - ob) > 0.01:
            breaks.append((rows[j - 1]['date'].date(), rows[j]['date'].date(),
                           round((nb - na) - ob, 2)))
    return breaks


def main():
    files = glob.glob(f"{BASE}/processed/*Chase*ALLREC*.csv")
    byacct = defaultdict(list)
    for f in files:
        a = acct(f)
        if a:
            byacct[a].append(f)

    wb = openpyxl.Workbook()
    idx = wb.active
    idx.title = "index"
    bold = Font(bold=True)
    hf = PatternFill("solid", fgColor="1F4E78")
    wfont = Font(bold=True, color="FFFFFF")

    idx.append(["Chase Master — one tab per bank account. Newest transactions at top. "
                "Stitched from overlapping Chase downloads."])
    idx.append([])
    idx.append(["account", "code", "name", "txns", "first", "last", "downloads used"])
    for c in range(1, 8):
        idx.cell(3, c).font = bold

    accts = ORDER + [a for a in sorted(byacct) if a not in ORDER]
    report = []
    for a in accts:
        if a not in byacct:
            continue
        code, name = ACCT_MAP.get(a, ("", ""))
        rows, windows = stitch(byacct[a])
        ws = wb.create_sheet(f"{a} {code} {name}".strip()[:31])
        for c, h in enumerate(HEAD, 1):
            x = ws.cell(1, c, h)
            x.font = wfont
            x.fill = hf
        for i, row in enumerate(rows, start=2):
            ws.cell(i, 1, row['details'])
            ws.cell(i, 2, row['date']).number_format = DATEF
            ws.cell(i, 3, row['desc'])
            ws.cell(i, 4, row['amt']).number_format = MONEY
            ws.cell(i, 5, row['typ'])
            if row['bal'] is not None:
                ws.cell(i, 6, row['bal']).number_format = MONEY
            ws.cell(i, 7, row['chk'])
            ws.cell(i, 8, row['src'])
        ws.freeze_panes = "A2"
        for col, w in {"A": 8, "B": 12, "C": 70, "D": 13, "E": 15,
                       "F": 14, "G": 11, "H": 30}.items():
            ws.column_dimensions[col].width = w
        idx.append([a, code, name, len(rows),
                    rows[-1]['date'].date().isoformat(),
                    rows[0]['date'].date().isoformat(),
                    ", ".join(w[0] for w in windows)])
        report.append((a, code, name, len(rows), qa_breaks(rows), windows))

    idx.freeze_panes = "A4"
    for col, w in {"A": 9, "B": 7, "C": 20, "D": 7, "E": 12, "F": 12, "G": 60}.items():
        idx.column_dimensions[col].width = w
    wb.save(OUT)

    print("wrote", OUT)
    for a, code, name, n, breaks, win in report:
        flag = "OK" if not breaks else f"*** {len(breaks)} BREAKS ***"
        print(f"  {a} {code} {name}: {n} txns  [{flag}]")
        for w in win:
            print(f"       {w[0]}  {w[1]} .. {w[2]}  ({w[3]})")
        for b in breaks[:5]:
            print(f"       break {b[0]}/{b[1]} off {b[2]}")


if __name__ == "__main__":
    main()
